#!/usr/bin/env python
import time
import openpyxl
import re
timeStart = time.time()

newlist = []

vocabulary = []

letter = {'a':1, 'b':2, 'c':3, 'd':4, 'e':5, 'f':6, 'g':7, 'h':8, 'i':9, 'j':10, 'k':11, 'l':12, 'm':13, 'n':14, 'o':15, 'p':16, 'q':17, 'r':18, 's':19, 't':20, 'u':21, 'v':22, 'w':23, 'x':24, 'y':25, 'z':26,'A':1, 'B':2, 'C':3, 'D':4, 'E':5, 'F':6, 'G':7, 'H':8, 'I':9, 'J':10, 'K':11, 'L':12, 'M':13, 'N':14, 'O':15, 'P':16, 'Q':17, 'R':18, 'S':19, 'T':20, 'U':21, 'V':22, 'W':23, 'X':24, 'Y':25, 'Z':26, '.':0, ' ':0, '-':0, '\'':0}

book = openpyxl.load_workbook('7000.xlsx')

sheetlist = book.sheetnames

for sheet_number in sheetlist:
    sheet = book[sheet_number]
    for number in range(1,sheet.max_row+1):
        if sheet.cell(row=number, column=1).value == None:
            pass
        else:
            regpressStr = re.match("^[^@]+", sheet.cell(row=number, column=1).value)
            newlist.append(regpressStr.group())

for spilt in newlist:
    #list長度是8799，但是從0~8798格才有value，還有重複的單字問題
    alphabet = list(spilt)
    alphabetlen = len(spilt)
    point = 0
    for total in range(0, alphabetlen):
        point += letter[alphabet[total]]
    if point == 100:
        vocabulary.append(spilt)
#list轉換成set，清除掉重複的元素
delduplicates = set(vocabulary)
print(delduplicates)
 
timeEnd = time.time()
print("It cost %f sec" % (timeEnd - timeStart))
