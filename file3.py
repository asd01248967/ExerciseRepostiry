#!/usr/bin/env python
# import time
# tStart = time.time()
import openpyxl
import re

newlist = []

match = []

letter = {'a':1, 'b':2, 'c':3, 'd':4, 'e':5, 'f':6, 'g':7, 'h':8, 'i':9, 'j':10, 'k':11, 'l':12, 'm':13, 'n':14, 'o':15, 'p':16, 'q':17, 'r':18, 's':19, 't':20, 'u':21, 'v':22, 'w':23, 'x':24, 'y':25, 'z':26,}

book = openpyxl.load_workbook('C:\\Users\\Jasper\\Desktop\\7000.xlsx')

sheetlist = book.sheetnames

for s_name in sheetlist:
    sheet = book[s_name]
    for number in range(1,sheet.max_row+1):
        if sheet.cell(row=number, column=1).value == None:
            pass
        else:
            rStr = re.match("^[^@]+", sheet.cell(row=number, column=1).value)
            print(rStr.group())
            print(type(rStr.group()))
            #newlist.append(rStr)


# for spilt in newlist:
    #list長度是8799，但是從0~8798格才有value，還有重複的單字問題
    # alphabet = list(spilt)
    # alphabetlen = len(spilt)
    # a = 0
    # print(spilt)

    # for total in range(0, alphabetlen):
    #     a += letter[alphabet[total]]
    # if a == 100:
    #     print(a)
        
#print(newlist[4657])

# tEnd = time.time()

# print("It cost %f sec" % (tEnd - tStart))
