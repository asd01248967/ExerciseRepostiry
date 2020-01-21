import openpyxl
import re

newlist = []

vocabulary = []

letter = {'a':1, 'b':2, 'c':3, 'd':4, 'e':5, 'f':6, 'g':7, 'h':8, 'i':9, 'j':10, 'k':11, 'l':12, 'm':13, 'n':14, 'o':15, 'p':16, 'q':17, 'r':18, 's':19, 't':20, 'u':21, 'v':22, 'w':23, 'x':24, 'y':25, 'z':26,'A':1, 'B':2, 'C':3, 'D':4, 'E':5, 'F':6, 'G':7, 'H':8, 'I':9, 'J':10, 'K':11, 'L':12, 'M':13, 'N':14, 'O':15, 'P':16, 'Q':17, 'R':18, 'S':19, 'T':20, 'U':21, 'V':22, 'W':23, 'X':24, 'Y':25, 'Z':26, '.':0, ' ':0, '-':0, '\'':0}

def consist():
    book = openpyxl.load_workbook('7000.xlsx')
    sheetlist = book.sheetnames
    for sheet_number in sheetlist:
        sheet = book[sheet_number]
        for number in range(1,sheet.max_row+1):
            if sheet.cell(row=number, column=1).value == None:
                pass
            else:
                regpressstr = re.match("^[^@]+", sheet.cell(row=number, column=1).value)
                newlist.append(regpressstr.group())
    for spilt in newlist:
        alphabet = list(spilt)
        point = 0
        for total in range(0, len(spilt)):
            point += letter[alphabet[total]]
        if point == 100:
            vocabulary.append(spilt)
    delduplicates = set(vocabulary)  #<----轉成set刪去重複單字時候才發生亂數排序的情況
    undolist = list(delduplicates)  #<----進行排序產生新的list並賦值給新變數
    sortlist = sorted(undolist)   #<----進行排序
    print(sortlist)

consist()


