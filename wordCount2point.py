#!/usr/bin/env python
import openpyxl
import re

newlist = []

vocabulary = []

letter = {'a':1, 'b':2, 'c':3, 'd':4, 'e':5, 'f':6, 'g':7, 'h':8, 'i':9, 'j':10, 'k':11, 'l':12, 'm':13, 'n':14, 'o':15, 'p':16, 'q':17, 'r':18, 's':19, 't':20, 'u':21, 'v':22, 'w':23, 'x':24, 'y':25, 'z':26,'A':1, 'B':2, 'C':3, 'D':4, 'E':5, 'F':6, 'G':7, 'H':8, 'I':9, 'J':10, 'K':11, 'L':12, 'M':13, 'N':14, 'O':15, 'P':16, 'Q':17, 'R':18, 'S':19, 'T':20, 'U':21, 'V':22, 'W':23, 'X':24, 'Y':25, 'Z':26, '.':0, ' ':0, '-':0, '\'':0}

def consist( requirements ):
    book = openpyxl.load_workbook('7000.xlsx')
    sheetlist = book.sheetnames
    for sheet_number in sheetlist:
        sheet = book[sheet_number]
        for number in range(1,sheet.max_row+1):
            if sheet.cell(row=number, column=1).value == None:
                pass
            else:
                regpressstr = re.match("^[^@]+", sheet.cell(row=number, column=1).value)
                newlist.append(regpressstr.group())
    for spilt in newlist:
        alphabet = list(spilt)
        point = 0
        for total in range(0, len(spilt)):
            point += letter[alphabet[total]]
        if point == int(requirements):
            vocabulary.append(spilt)
    delduplicates = set(vocabulary)  #<----轉成set刪去重複單字時候才發生亂數排序的情況
    undolist = list(delduplicates)  #<----產生新的list並賦值給新變數
    sortlist = sorted(undolist)   #<----進行排序
    print(sortlist)

def printme( name ):
    userinput = list(name)
    record = 0
    for firstword in range(0, len(userinput)):
        try:
            record += letter[userinput[firstword]]
        except Exception as e:
            print("your input", e, "is a not Qualification Character Symbol that can't count point")
    print(name, "is point get total :",record)

def main():
    while 1:
        print("*********************************")
        print("choosen one you wanna to execute.")
        print("1.Pickup a word from file thath qualified convert point special point on user ")
        print("2.Import and process 7000's word and show requirements special qualified point's word in file")
        print("3.Exit ")
        choosen = input("which number is you want to exect : ")
        try:
            if int(choosen) == 1:
                name = str(input("Enter Identify Word: "))
                printme(name)
            elif int(choosen) == 2:
                requirements = input("requirements special point is :")
                consist(requirements)
            elif int(choosen) == 3:
                break
            else:
                print("choosen neither 1 nor 2, please choosen again")
        except Exception as e:
            print(e, ",This filed restricttyped to only accepted numeric")

main()